# -*- coding: utf-8 -*-
"""업체 연락처 수집기 (여우알바 + 퀸알바).

고객 PC에서 실행하는 단일 GUI 도구.
1) 각 사이트에서 본인인증(비회원 인증)을 직접 완료한 뒤
2) 버튼을 누르면 로그인된 세션 쿠키로 모든 상세 광고를 자동 수집하고
3) 바탕화면에 엑셀(업체명/담당자/전화번호)을 저장한다.

여우알바 상세표는 '담당자' 헤더 이후 <font color="#525252"> 값들이
[업종, 담당자, 상호(업체명), 주소, 연락처(전화), 카톡] 순서(고객 샘플 80/80 검증).
퀸알바는 라벨 셀(>상호<, >담당자<, >전화번호<) 다음 첫 비어있지 않은 <td> 값.
"""
import io
import json
import os
import platform
import re
import struct
import sys
import threading
import time
import traceback
import zipfile

# --- 한글 print 안전 (cp1252 콘솔에서 UnicodeEncodeError 방지) ---
for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

import requests

# --- 실행 모드 플래그 ---
EXTRACTOR_AUTO = os.environ.get("EXTRACTOR_AUTO") == "1"
EXTRACTOR_HEADLESS = os.environ.get("EXTRACTOR_HEADLESS") == "1"
EXTRACTOR_MAX = int(os.environ.get("EXTRACTOR_MAX", "0"))  # 0=전체, N=상세 N건 제한
# 우리(개발자) 전용 진단 모드: 실제 고객 진입점은 항상 GUI 다.
EXTRACTOR_DRIVERTEST = os.environ.get("EXTRACTOR_DRIVERTEST") == "1"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
MAX_PAGES = 400  # 무한 루프 방지 상한

# ===================== 앱 상수 / 진단 =====================
APP_NAME = "업체연락처수집기"
APP_VERSION = "1.2.0"
CUSTOMER_ID = "244448"                       # Kmong partnerId (WCompany)
WORKS_API = "https://works.insu.ng/works/api"
ARTIFACT_SOURCE = "wcompany-collector-diag"

RUN_ID = time.strftime("%Y%m%d-%H%M%S")
DIAG = {}                 # 구조화 진단 값 (chrome/driver/steps)
SNAPSHOTS = {}            # 파일명 -> bytes (페이지 HTML 등, 리포트 ZIP 에 동봉)
_LOG_LINES = []
_LOG_LOCK = threading.Lock()
_MAX_LOG_LINES = 20000


def app_data_dir():
    """로그·드라이버 캐시를 두는 폴더. 경로에 한글이 섞이지 않게 ASCII 이름을 쓴다."""
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    d = os.path.join(base, "WCompanyCollector")
    try:
        os.makedirs(d, exist_ok=True)
    except Exception:
        d = os.path.expanduser("~")
    return d


def _log_path():
    # --windowed exe 는 stdout 이 없다. CI/지원 목적으로 로그 경로를 고정할 수 있게 한다.
    override = os.environ.get("EXTRACTOR_LOG_FILE")
    if override:
        try:
            parent = os.path.dirname(os.path.abspath(override))
            if parent:
                os.makedirs(parent, exist_ok=True)
        except Exception:
            pass
        return override
    d = os.path.join(app_data_dir(), "logs")
    try:
        os.makedirs(d, exist_ok=True)
    except Exception:
        return os.path.join(app_data_dir(), "run-%s.log" % RUN_ID)
    return os.path.join(d, "run-%s-%s.log" % (CUSTOMER_ID, RUN_ID))


LOG_PATH = _log_path()


def record(msg):
    """화면에 띄우지 않고 로그 파일과 메모리에만 남긴다(스택트레이스 등)."""
    line = "%s  %s" % (time.strftime("%H:%M:%S"), msg)
    try:
        with _LOG_LOCK:
            _LOG_LINES.append(line)
            if len(_LOG_LINES) > _MAX_LOG_LINES:
                del _LOG_LINES[:len(_LOG_LINES) - _MAX_LOG_LINES]
            with open(LOG_PATH, "a", encoding="utf-8") as f:
                f.write(line + "\n")
    except Exception:
        pass


def snapshot(name, blob):
    """페이지 HTML 등 증거를 리포트 ZIP 에 담기 위해 보관(용량 제한)."""
    try:
        if isinstance(blob, str):
            blob = blob.encode("utf-8", "replace")
        if len(SNAPSHOTS) >= 8:
            return
        SNAPSHOTS[name] = blob[:400_000]
    except Exception:
        pass


def _log_text():
    with _LOG_LOCK:
        lines = list(_LOG_LINES)
    if len(lines) > 3000:      # 가운데를 잘라 앞뒤를 남긴다
        lines = lines[:1500] + ["... (중략 %d줄) ..." % (len(lines) - 3000)] + lines[-1500:]
    return "\n".join(lines)


def _diagnostics(kind):
    return {
        "app": APP_NAME, "version": APP_VERSION, "customerId": CUSTOMER_ID,
        "runId": RUN_ID, "kind": kind,
        "os": platform.platform(), "python": platform.python_version(),
        "frozen": bool(getattr(sys, "frozen", False)),
        "arch": platform.machine(), "logPath": LOG_PATH,
        "flags": {"auto": EXTRACTOR_AUTO, "headless": EXTRACTOR_HEADLESS,
                  "max": EXTRACTOR_MAX, "drivertest": EXTRACTOR_DRIVERTEST},
        "diag": DIAG,
    }


def report(kind, summary="", wait=False):
    """Artifacts API 로 이번 실행의 로그/진단/페이지를 올린다.
    어떤 실패도 프로그램을 멈추지 않는다(전부 catch-all)."""
    def task():
        try:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr("run-%s.log" % RUN_ID, _log_text())
                z.writestr("diagnostics.json",
                           json.dumps(_diagnostics(kind), ensure_ascii=False, indent=2, default=str))
                for name, blob in list(SNAPSHOTS.items()):
                    try:
                        z.writestr(name, blob)
                    except Exception:
                        pass
            text = ("[%s v%s] customer=%s run=%s kind=%s\n%s\n"
                    "chrome=%s\ndriver=%s\nos=%s python=%s frozen=%s\nlog=%s"
                    % (APP_NAME, APP_VERSION, CUSTOMER_ID, RUN_ID, kind, summary,
                       DIAG.get("chrome"), DIAG.get("driver"), platform.platform(),
                       platform.python_version(), bool(getattr(sys, "frozen", False)), LOG_PATH))
            fname = "%s-%s-%s.zip" % (CUSTOMER_ID, kind, RUN_ID)
            r = requests.post(
                WORKS_API,
                data={"customerId": CUSTOMER_ID, "source": ARTIFACT_SOURCE, "text": text[:6000]},
                files={"file": (fname, buf.getvalue(), "application/zip")},
                timeout=40,
            )
            record("[진단전송] %s -> HTTP %s %s" % (kind, r.status_code, r.text[:300]))
        except Exception as e:
            record("[진단전송 실패] %s: %s" % (kind, e))

    try:
        t = threading.Thread(target=task, daemon=True)
        t.start()
        if wait:
            t.join(60)
        return t
    except Exception:
        return None

# ===================== 공통 파서 유틸 =====================
VAL_RE = re.compile(r'<font color="#525252">(.*?)</font>', re.S | re.I)
SP = lambda w: r'\s*'.join(map(re.escape, w))
CLEAN = lambda s: re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', s).replace('&nbsp;', ' ')).strip()
PHONE_RE = re.compile(r'0\d{1,2}[-. )]\d{3,4}[-. ]\d{4}')


def parse_fox(html):
    m = re.search(SP('담당자'), html)
    if not m:
        return None
    vals = [CLEAN(mm.group(1)) for mm in VAL_RE.finditer(html) if mm.start() > m.start()]
    if len(vals) < 5:
        return None
    phone = vals[4]
    if not PHONE_RE.fullmatch(phone):
        pm = PHONE_RE.search(' '.join(vals[3:7]))
        phone = pm.group(0) if pm else phone
    if not PHONE_RE.search(phone):
        return None
    return {
        '업체명': vals[2], '담당자': vals[1],
        '전화번호': re.sub(r'[ .)]', '-', phone).strip('-'),
    }


def field_queen(label, html):
    m = re.search(r'>\s*' + re.escape(label) + r'\s*<', html)
    if not m:
        return ''
    for c in re.findall(r'<td[^>]*>(.*?)</td>', html[m.end():m.end() + 500], re.S | re.I):
        v = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', c).replace('&nbsp;', ' ')).strip()
        if v:
            return v
    return ''


def parse_queen(html):
    name = field_queen('상호', html)
    mgr = field_queen('담당자', html)
    phone = field_queen('전화번호', html)
    pm = PHONE_RE.search(phone)
    if not pm:
        return None
    return {
        '업체명': name, '담당자': mgr,
        '전화번호': re.sub(r'[ .)]', '-', pm.group(0)).strip('-'),
    }


# ===================== 사이트 URL =====================
# 실측: foxalba 의 실제 경로는 /offer/ 이며 /job/ 은 404.
FOX_BASE = "https://www.foxalba.com"
FOX_LIST = FOX_BASE + "/offer/offer_jobpart.asp?page={page}"
FOX_DETAIL = FOX_BASE + "/offer/offer_content.asp?idx={idx}"
FOX_DETAIL_RE = re.compile(r'offer_content\.asp\?idx=([0-9]+)', re.I)

QUEEN_BASE = "https://www.queenalba.net"
# 본인인증(비회원 인증)은 메인 페이지(adult_index.php)에서 진행한다.
# 실측: 목록/상세는 루트 경로다. /guin/ 하위는 존재하지 않아 error.php 로 튕긴다.
#   O  https://www.queenalba.net/guin_list.php?pg=1     (인증 후 1.7MB, 상세링크 700+)
#   X  https://www.queenalba.net/guin/guin_list.php     (-> error.php, 0건)
# 상세 링크는 HTML 에서 &amp; 로 인코딩되고 cou= 가 비어있는 경우가 많아 num 만 신뢰한다.
QUEEN_ENTRY = QUEEN_BASE + "/"
QUEEN_GUIN_FIRST = QUEEN_BASE + "/guin_list.php?pg=1"
QUEEN_LIST = QUEEN_BASE + "/guin_list.php?pg={pg}"
QUEEN_DETAIL = QUEEN_BASE + "/guin_detail.php?num={num}"
QUEEN_DETAIL_RE = re.compile(r'guin_detail\.php\?num=([0-9]+)', re.I)


# ===================== Selenium 드라이버 =====================
def make_driver(headless):
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--window-size=1400,1000")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-popup-blocking")  # 퀸알바 본인인증 팝업 허용
    opts.add_argument("--disable-features=IsolateOrigins,site-per-process")
    opts.add_argument(f"--user-agent={UA}")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    drv = webdriver.Chrome(options=opts)  # Selenium Manager 가 드라이버 자동 설치
    # 퀸알바 봇 탐지 우회: navigator.webdriver 등 자동화 흔적을 새 문서마다 덮어쓴다.
    try:
        drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
Object.defineProperty(navigator, 'languages', {get: () => ['ko-KR', 'ko', 'en-US', 'en']});
"""
        })
    except Exception:
        pass
    try:
        drv.set_page_load_timeout(45)
    except Exception:
        pass
    return drv


class ChromeNotFound(Exception):
    """이 PC 에 Chrome 이 설치되어 있지 않다."""


class DriverStartError(Exception):
    """모든 드라이버 시도가 실패했다."""


# ---------- 설치된 Chrome 버전 탐지 ----------
def _win_file_version(path):
    """chrome.exe 의 파일 버전('151.0.7922.138')을 ctypes 로 직접 읽는다.
    subprocess 를 쓰지 않으므로 --noconsole exe 에서 콘솔 창이 깜빡이지 않고,
    PowerShell 경로 이스케이프 문제도 없다."""
    try:
        import ctypes
        from ctypes import wintypes
        ver = ctypes.WinDLL("version.dll")
        ver.GetFileVersionInfoSizeW.argtypes = [wintypes.LPCWSTR, ctypes.POINTER(wintypes.DWORD)]
        ver.GetFileVersionInfoSizeW.restype = wintypes.DWORD
        ver.GetFileVersionInfoW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD,
                                            wintypes.DWORD, ctypes.c_void_p]
        ver.GetFileVersionInfoW.restype = wintypes.BOOL
        ver.VerQueryValueW.argtypes = [ctypes.c_void_p, wintypes.LPCWSTR,
                                       ctypes.POINTER(ctypes.c_void_p), ctypes.POINTER(ctypes.c_uint)]
        ver.VerQueryValueW.restype = wintypes.BOOL

        unused = wintypes.DWORD(0)
        size = ver.GetFileVersionInfoSizeW(path, ctypes.byref(unused))
        if not size:
            return None
        buf = ctypes.create_string_buffer(size)
        if not ver.GetFileVersionInfoW(path, 0, size, buf):
            return None
        ptr, ln = ctypes.c_void_p(), ctypes.c_uint()
        if not ver.VerQueryValueW(buf, "\\", ctypes.byref(ptr), ctypes.byref(ln)):
            return None
        data = ctypes.string_at(ptr, ln.value)
        sig, _struc, ms, ls = struct.unpack("<4I", data[:16])
        if sig != 0xFEEF04BD:
            return None
        return "%d.%d.%d.%d" % (ms >> 16, ms & 0xFFFF, ls >> 16, ls & 0xFFFF)
    except Exception:
        return None


def _registry_chrome_versions():
    """HKCU/HKLM 의 BLBeacon·Google Update 항목에서 Chrome 버전을 읽는다."""
    out = []
    try:
        import winreg
    except Exception:
        return out
    guid = r"{8A69D345-D564-463c-AFF1-A69D9E530F96}"
    keys = [
        (winreg.HKEY_CURRENT_USER, r"Software\Google\Chrome\BLBeacon", "version"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Google\Chrome\BLBeacon", "version"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Wow6432Node\Google\Chrome\BLBeacon", "version"),
        (winreg.HKEY_CURRENT_USER, r"Software\Google\Update\Clients" + "\\" + guid, "pv"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Google\Update\Clients" + "\\" + guid, "pv"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Wow6432Node\Google\Update\Clients" + "\\" + guid, "pv"),
    ]
    wow_flags = [0, getattr(winreg, "KEY_WOW64_64KEY", 0), getattr(winreg, "KEY_WOW64_32KEY", 0)]
    for hive, sub, name in keys:
        for flag in wow_flags:
            try:
                with winreg.OpenKey(hive, sub, 0, winreg.KEY_READ | flag) as k:
                    v, _ = winreg.QueryValueEx(k, name)
                if v:
                    out.append(("reg:%s\\%s" % (sub, name), str(v)))
                break
            except Exception:
                continue
    return out


def _chrome_exe_candidates():
    """설치된 chrome.exe 후보 경로(실제로 존재하는 것만)."""
    cands = []
    if sys.platform.startswith("win"):
        try:
            import winreg
            for hive in (winreg.HKEY_CURRENT_USER, winreg.HKEY_LOCAL_MACHINE):
                try:
                    key = r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"
                    with winreg.OpenKey(hive, key) as k:
                        v, _ = winreg.QueryValueEx(k, "")
                    if v:
                        cands.append(str(v).strip('"'))
                except Exception:
                    continue
        except Exception:
            pass
        for env in ("PROGRAMFILES", "PROGRAMW6432", "PROGRAMFILES(X86)", "LOCALAPPDATA"):
            base = os.environ.get(env)
            if base:
                cands.append(os.path.join(base, "Google", "Chrome", "Application", "chrome.exe"))
    else:
        cands += ["/usr/bin/google-chrome", "/usr/bin/google-chrome-stable",
                  "/usr/bin/chromium", "/usr/bin/chromium-browser"]
    out, seen = [], set()
    for c in cands:
        try:
            key = os.path.normpath(c).lower()
            if key in seen:
                continue
            seen.add(key)
            if os.path.isfile(c):
                out.append(c)
        except Exception:
            continue
    return out


def _versioned_dir_version(exe_path):
    """Chrome 설치 폴더의 '151.0.7922.138' 형태 하위 디렉터리에서 최신 버전을 읽는다."""
    try:
        d = os.path.dirname(exe_path)
        best = None
        for n in os.listdir(d):
            if re.fullmatch(r"\d+\.\d+\.\d+\.\d+", n) and os.path.isdir(os.path.join(d, n)):
                t = tuple(int(x) for x in n.split("."))
                if best is None or t > best[0]:
                    best = (t, n)
        return best[1] if best else None
    except Exception:
        return None


def _posix_chrome_version(exe_path):
    if sys.platform.startswith("win"):
        return None
    try:
        import subprocess
        r = subprocess.run([exe_path, "--version"], capture_output=True, text=True, timeout=15)
        m = re.search(r"(\d+\.\d+\.\d+\.\d+)", (r.stdout or "") + (r.stderr or ""))
        return m.group(1) if m else None
    except Exception:
        return None


def detect_chrome():
    """설치된 Chrome 의 경로와 버전을 여러 독립 신호로 교차 확인한다.

    우선순위는 '실제로 실행될 chrome.exe 의 파일 버전' -> 설치 폴더의 버전 디렉터리
    -> 레지스트리. 레지스트리(BLBeacon)는 업데이트 대기 상태에서 실제 실행 파일보다
    앞서 갱신되는 경우가 있어 파일 버전을 더 신뢰한다."""
    signals, exe = [], None
    for p in _chrome_exe_candidates():
        fv = _win_file_version(p) or _posix_chrome_version(p)
        if fv:
            signals.append(("file:%s" % p, fv))
            exe = exe or p
        dv = _versioned_dir_version(p)
        if dv:
            signals.append(("dir:%s" % os.path.dirname(p), dv))
            exe = exe or p
    cands = _chrome_exe_candidates()
    if exe is None and cands:
        exe = cands[0]
    signals.extend(_registry_chrome_versions())

    major = full = None
    for _src, v in signals:
        m = re.match(r"\s*(\d+)\.", str(v))
        if m:
            major, full = int(m.group(1)), str(v).strip()
            break
    return {"exe": exe, "major": major, "full": full, "signals": signals}


# ---------- 브라우저 버전에 정확히 맞는 chromedriver 확보 ----------
CFT_MILESTONES_URL = ("https://googlechromelabs.github.io/chrome-for-testing/"
                      "latest-versions-per-milestone-with-downloads.json")


def _driver_platform():
    if sys.platform.startswith("win"):
        return "win64" if sys.maxsize > 2 ** 32 else "win32"
    if sys.platform == "darwin":
        return "mac-arm64" if platform.machine().lower() in ("arm64", "aarch64") else "mac-x64"
    return "linux64"


def fetch_chromedriver_for_major(major, log):
    """Chrome for Testing 에서 major 에 정확히 대응하는 chromedriver 를 받아 캐시한다.
    이것이 버전 불일치를 구조적으로 막는 핵심: 최신 드라이버를 받지 않고
    '설치된 브라우저의 메이저 버전' 드라이버만 받는다."""
    exe_name = "chromedriver.exe" if sys.platform.startswith("win") else "chromedriver"
    ddir = os.path.join(app_data_dir(), "drivers", str(major))
    dpath = os.path.join(ddir, exe_name)
    if os.path.isfile(dpath) and os.path.getsize(dpath) > 1_000_000:
        log("  드라이버 캐시 사용: %s" % dpath)
        return dpath

    r = requests.get(CFT_MILESTONES_URL, timeout=30)
    r.raise_for_status()
    entry = (r.json().get("milestones") or {}).get(str(major))
    if not entry:
        raise RuntimeError("chrome-for-testing 에 milestone %s 가 없습니다" % major)
    plat = _driver_platform()
    downloads = (entry.get("downloads") or {}).get("chromedriver") or []
    url = None
    for want in (plat, "win32", "win64"):
        for it in downloads:
            if it.get("platform") == want:
                url = it.get("url")
                break
        if url:
            break
    if not url:
        raise RuntimeError("milestone %s 에 %s 용 chromedriver 가 없습니다" % (major, plat))

    log("  chromedriver %s (%s) 내려받는 중..." % (entry.get("version"), plat))
    zr = requests.get(url, timeout=180)
    zr.raise_for_status()
    os.makedirs(ddir, exist_ok=True)
    with zipfile.ZipFile(io.BytesIO(zr.content)) as z:
        member = next((n for n in z.namelist()
                       if n == exe_name or n.endswith("/" + exe_name)), None)
        if member is None:
            raise RuntimeError("zip 안에서 %s 를 찾지 못했습니다" % exe_name)
        with z.open(member) as src, open(dpath, "wb") as dst:
            dst.write(src.read())
    try:
        os.chmod(dpath, 0o755)
    except Exception:
        pass
    DIAG.setdefault("driverDownloads", []).append(
        {"major": major, "version": entry.get("version"), "platform": plat, "url": url})
    log("  드라이버 준비 완료: chromedriver %s" % entry.get("version"))
    return dpath


def _selenium_manager_driver(chrome_exe, log):
    """Selenium Manager 가 '설치된 브라우저에 맞는' 드라이버를 고르게 한다.
    최신 버전을 받는 것이 아니라 브라우저에 맞춰 해석하므로 마지막 안전망으로 적합하다."""
    from selenium.webdriver.chrome.options import Options as ChromeOpts
    from selenium.webdriver.chrome.service import Service as ChromeService
    o = ChromeOpts()
    if chrome_exe:
        o.binary_location = chrome_exe
    svc = ChromeService()
    from selenium.webdriver.common.driver_finder import DriverFinder
    try:
        path = DriverFinder(svc, o).get_driver_path()          # selenium 4.2x+
    except TypeError:
        path = DriverFinder.get_path(svc, o)                   # 구버전 호환
    log("  Selenium Manager 드라이버: %s" % path)
    return path


def _queen_options():
    """시도마다 새로 만든다. 기동에 실패한 ChromeOptions 는 재사용할 수 없다."""
    import undetected_chromedriver as uc
    opts = uc.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--window-size=1400,1000")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--lang=ko-KR")
    if EXTRACTOR_HEADLESS:
        opts.add_argument("--headless=new")
    return opts


def make_uc_driver(log=None):
    """퀸알바 전용 - undetected_chromedriver 로 봇 감지 우회.

    버전 불일치(드라이버가 브라우저보다 최신)는 이 고객 환경의 상시 조건이므로,
    '설치된 Chrome 의 메이저 버전에 맞는 드라이버'를 먼저 직접 받아 쓴다.
    어떤 시도도 '최신 드라이버 자동 다운로드'로 끝나지 않는다."""
    import undetected_chromedriver as uc
    log = log or record

    info = detect_chrome()
    DIAG["chrome"] = {"exe": info["exe"], "major": info["major"], "full": info["full"]}
    DIAG["chromeSignals"] = ["%s=%s" % (s, v) for s, v in info["signals"]]
    for s, v in info["signals"]:
        record("  Chrome 신호: %s -> %s" % (s, v))
    if info["exe"] is None and info["major"] is None:
        raise ChromeNotFound()

    major = info["major"]
    log("설치된 Chrome: %s (major=%s)" % (info["full"] or "확인 불가", major))

    attempts = []
    if major:
        attempts.append(("Chrome %d 에 맞춘 드라이버(직접 지정)" % major, {"pin": major}))
        attempts.append(("uc 다운로드 version_main=%d" % major, {"vm": major}))
        attempts.append(("uc 다운로드 version_main=%d" % (major - 1), {"vm": major - 1}))
        attempts.append(("uc 다운로드 version_main=%d" % (major + 1), {"vm": major + 1}))
        attempts.append(("Chrome %d 에 맞춘 드라이버(직접 지정)" % (major - 1), {"pin": major - 1}))
    # 마지막 안전망도 '최신 자동 다운로드'가 아니라 설치된 브라우저 기준 해석이다.
    attempts.append(("Selenium Manager 가 고른 드라이버", {"sm": True}))

    errors = []
    for label, spec in attempts:
        drv = None
        try:
            log("드라이버 시도: %s" % label)
            kw = {"options": _queen_options(), "use_subprocess": True}
            if info["exe"]:
                kw["browser_executable_path"] = info["exe"]
            if "pin" in spec:
                kw["driver_executable_path"] = fetch_chromedriver_for_major(spec["pin"], log)
            elif "vm" in spec:
                kw["version_main"] = spec["vm"]
            elif spec.get("sm"):
                kw["driver_executable_path"] = _selenium_manager_driver(info["exe"], log)
            drv = uc.Chrome(**kw)
            caps = {}
            try:
                caps = drv.capabilities or {}
            except Exception:
                pass
            cd = (caps.get("chrome") or {}).get("chromedriverVersion", "?")
            bv = caps.get("browserVersion", "?")
            DIAG["driver"] = {"attempt": label, "browserVersion": bv,
                              "chromedriverVersion": cd, "failedAttempts": errors}
            log("드라이버 연결 성공: %s" % label)
            log("  브라우저 %s / chromedriver %s" % (bv, cd))
            try:
                drv.set_page_load_timeout(45)
            except Exception:
                pass
            return drv
        except Exception as e:
            first = str(e).strip().splitlines()[0] if str(e).strip() else e.__class__.__name__
            errors.append("%s -> %s" % (label, first[:300]))
            log("  실패: %s" % first[:200])
            record("  [상세] %s" % traceback.format_exc())
            try:
                if drv is not None:
                    drv.quit()
            except Exception:
                pass

    DIAG["driver"] = {"attempt": None, "failedAttempts": errors}
    raise DriverStartError(" | ".join(errors))


def friendly_error(e):
    """고객 화면에는 스택트레이스 대신 이 한 줄만 보여준다."""
    if isinstance(e, ChromeNotFound):
        return ("이 PC에서 Chrome 브라우저를 찾지 못했습니다. "
                "Chrome을 설치한 뒤 다시 실행해 주세요.")
    s = str(e)
    low = s.lower()
    if isinstance(e, DriverStartError) or "session not created" in low or "only supports chrome" in low:
        return ("Chrome 연결에 실패했습니다. 열려 있는 Chrome 창을 모두 닫고 다시 눌러주세요. "
                "계속 같은 증상이면 자동으로 전송된 진단 기록을 보고 바로 조치하겠습니다.")
    if "cannot connect to chrome" in low or "chrome not reachable" in low:
        return "Chrome 실행에 실패했습니다. 열려 있는 Chrome 창을 모두 닫고 다시 눌러주세요."
    if "timed out" in low or "timeout" in low:
        return "응답이 느려 시간이 초과되었습니다. 잠시 후 다시 시도해 주세요."
    if "connection" in low or "network" in low or "resolve" in low:
        return "인터넷 연결이 불안정합니다. 연결을 확인한 뒤 다시 시도해 주세요."
    return "작업 중 문제가 발생했습니다. 진단 기록을 전송했으니 확인 후 조치하겠습니다."


# ===================== 자동 업데이트 =====================
# 파일명 규약: 새 빌드는 항상 버전 접미사가 붙은 파일명으로만 배포하고 version 파일의
# exeUrl 이 그 새 경로를 가리킨다. 이미 서빙 중인 파일명을 덮어쓰면 Cloudflare 엣지
# 캐시가 옛 바이트를 계속 내보내 재시작 루프가 생긴다.
UPDATE_VERSION_URL = "https://works.insu.ng/works/public/%s/version-wcompany-collector.json" % CUSTOMER_ID
MIN_EXE_BYTES = 5_000_000


def _version_tuple(v):
    try:
        return tuple(int(x) for x in str(v).strip().split("."))
    except Exception:
        return (0,)


def check_update(log, is_busy):
    """시작 직후 한 번만 확인한다.

    이 프로그램은 고객이 브라우저 본인인증을 직접 하는 단계형 도구라,
    수집이 시작된 뒤 재시작하면 진행 중인 작업이 날아간다. 그래서 아직 아무 것도
    시작하지 않은 시점에만 교체하고, 그 외에는 안내만 남긴다."""
    try:
        r = requests.get(UPDATE_VERSION_URL, timeout=10, headers={"Cache-Control": "no-cache"})
        if r.status_code != 200:
            return
        data = r.json()
        latest, exe_url = str(data.get("version", "")).strip(), data.get("exeUrl")
        if not latest or not exe_url:
            return
        if _version_tuple(latest) <= _version_tuple(APP_VERSION):
            record("업데이트 확인: 최신 버전(%s) 사용 중" % APP_VERSION)
            return
        if not getattr(sys, "frozen", False):
            record("업데이트 있음(%s) - 개발 실행이라 교체 생략" % latest)
            return
        if is_busy():
            log("새 버전(%s)이 있습니다. 이번 작업을 끝내고 프로그램을 다시 켜주세요." % latest)
            return

        log("새 버전 %s 을(를) 내려받는 중입니다..." % latest)
        import tempfile
        fd, tmp = tempfile.mkstemp(suffix=".exe")
        os.close(fd)
        total = 0
        with requests.get(exe_url, timeout=180, stream=True,
                          headers={"Cache-Control": "no-cache"}) as resp:
            if resp.status_code != 200:
                os.unlink(tmp)
                return
            expected = resp.headers.get("Content-Length")
            with open(tmp, "wb") as f:
                for chunk in resp.iter_content(chunk_size=1 << 16):
                    if chunk:
                        f.write(chunk)
                        total += len(chunk)
        if (expected and expected.isdigit() and total != int(expected)) or total < MIN_EXE_BYTES:
            record("업데이트 다운로드 불완전: 받음=%s 기대=%s" % (total, expected))
            os.unlink(tmp)
            return

        current, pid = sys.executable, os.getpid()
        script = (
            "@echo off\r\n:wait\r\n"
            'tasklist /FI "PID eq %d" 2>NUL | find "%d" >NUL\r\n' % (pid, pid) +
            "if not errorlevel 1 (\r\n  timeout /t 1 /nobreak >NUL\r\n  goto wait\r\n)\r\n"
            'copy /y "%s" "%s" >NUL\r\n' % (tmp, current) +
            'start "" "%s"\r\n' % current +
            'del "%~f0"\r\n'
        )
        fd, bat = tempfile.mkstemp(suffix=".bat")
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(script)
        log("새 버전으로 다시 시작합니다...")
        report("update-restart", summary="%s -> %s" % (APP_VERSION, latest), wait=True)
        import subprocess
        subprocess.Popen(["cmd.exe", "/c", bat],
                         creationflags=getattr(subprocess, "DETACHED_PROCESS", 0)
                         | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0))
        os._exit(0)
    except Exception as e:
        record("업데이트 확인 실패(무시): %s" % e)


def session_from_driver(driver):
    """로그인된 Chrome 세션의 쿠키를 requests.Session 으로 옮긴다(상세 페이지 고속 수집)."""
    sess = requests.Session()
    sess.headers.update({"User-Agent": UA})
    try:
        for c in driver.get_cookies():
            try:
                sess.cookies.set(c["name"], c["value"], domain=c.get("domain"), path=c.get("path", "/"))
            except Exception:
                sess.cookies.set(c["name"], c["value"])
    except Exception:
        pass
    return sess


def _get(sess, url, encoding=None, referer=None):
    headers = {"Referer": referer} if referer else {}
    r = sess.get(url, headers=headers, timeout=30)
    if encoding:
        r.encoding = encoding
    return r.text


# ===================== 여우알바 수집 =====================
def scrape_fox(sess, log, max_items=0):
    idxs = []
    seen = set()
    for page in range(1, MAX_PAGES + 1):
        try:
            html = _get(sess, FOX_LIST.format(page=page), encoding="euc-kr")
        except Exception as e:
            log(f"  목록 {page}페이지 오류: {e}")
            break
        found = FOX_DETAIL_RE.findall(html)
        new = [i for i in found if i not in seen]
        for i in found:
            seen.add(i)
        log(f"페이지 {page}... 상세링크 {len(found)}개 (신규 {len(new)}개)")
        if not found:
            break
        if not new and page > 1:
            break
        idxs.extend(new)
        if max_items and len(idxs) >= max_items:
            idxs = idxs[:max_items]
            break
        time.sleep(0.2)

    rows, phones = [], set()
    total = len(idxs)
    for n, idx in enumerate(idxs, 1):
        try:
            html = _get(sess, FOX_DETAIL.format(idx=idx), encoding="euc-kr",
                        referer=FOX_LIST.format(page=1))
        except Exception:
            continue
        r = parse_fox(html)
        if not r:
            continue
        if r['전화번호'] in phones:
            continue
        phones.add(r['전화번호'])
        rows.append(r)
        if n % 5 == 0 or n == total:
            log(f"연락처 수집 중 ({len(rows)}건)... {n}/{total}")
        time.sleep(0.05)
    log(f"여우알바 수집 완료: {len(rows)}건")
    return rows


# ===================== 퀸알바 수집 =====================
def queen_list_ok(sess, log):
    """인증 세션 점검: guin_list 1페이지에서 상세 num 링크가 보이는지 확인.
    미인증/세션없음이면 adult_index.php 또는 error.php 로 튕긴다."""
    url = QUEEN_LIST.format(pg=1)
    try:
        r = sess.get(url, timeout=30)
        r.encoding = "utf-8"
    except Exception as e:
        log(f"  [퀸] 목록 점검 요청 실패: {e}")
        return False, 0
    bounced = ("adult_index" in r.url) or ("error.php" in r.url)
    nums = set(QUEEN_DETAIL_RE.findall(r.text))
    note = "  ← 인증세션 없음(adult/error 로 리다이렉트)" if bounced else ""
    log(f"  [퀸] 목록 점검: final={r.url} size={len(r.text)} 상세 num={len(nums)}{note}")
    # 실패를 재현할 수 있도록 이 시점의 페이지와 응답 정보를 증거로 남긴다.
    snapshot("queen-list-check.html", r.text)
    DIAG["queenListCheck"] = {"finalUrl": r.url, "status": r.status_code,
                              "size": len(r.text), "detailNums": len(nums),
                              "bounced": bounced,
                              "cookies": sorted(sess.cookies.get_dict().keys())}
    return (len(nums) > 0 and not bounced), len(nums)


def scrape_queen(sess, log, max_items=0):
    """퀸알바 수집 - 인증된 쿠키를 옮긴 requests 세션으로 고속 수집(여우와 동일 방식).
    상세 페이지는 마감 공고에서 JS 알럿이 떠 Selenium 직접 탐색은 위험 → requests 사용."""
    # 1) 목록 순회: guin_detail 의 num 수집(중복 제거)
    nums, seen = [], set()
    empty_streak = 0
    for pg in range(1, MAX_PAGES + 1):
        url = QUEEN_LIST.format(pg=pg)
        try:
            r = sess.get(url, timeout=30)
            r.encoding = "utf-8"
        except Exception as e:
            log(f"  목록 {pg}페이지 오류: {e}")
            break
        if "adult_index" in r.url or "error.php" in r.url:
            log(f"  목록 {pg}페이지: 인증 세션이 풀려 {r.url} 로 이동됨 — 목록 순회 중단")
            break
        page_nums = list(dict.fromkeys(QUEEN_DETAIL_RE.findall(r.text)))
        new = [n for n in page_nums if n not in seen]
        for n in page_nums:
            seen.add(n)
        nums.extend(new)
        log(f"페이지 {pg}... size={len(r.text)} 상세 num {len(page_nums)}개 (신규 {len(new)}개, 누적 {len(nums)})")
        if not page_nums:
            empty_streak += 1
            if empty_streak >= 2:
                log(f"  -> {pg}페이지 연속 빈 목록, 순회 종료")
                break
        else:
            empty_streak = 0
        if not new and pg > 1:
            log(f"  -> {pg}페이지부터 신규 num 없음, 순회 종료")
            break
        if max_items and len(nums) >= max_items:
            nums = nums[:max_items]
            break
        time.sleep(0.2)

    # 2) 상세 페이지 수집 → 업체명/담당자/전화번호
    rows, phones = [], set()
    closed = skipped = 0
    total = len(nums)
    log(f"  상세(업체) 페이지 {total}개 수집 시작...")
    for i, num in enumerate(nums, 1):
        try:
            r = sess.get(QUEEN_DETAIL.format(num=num), timeout=30)
            r.encoding = "utf-8"
        except Exception:
            skipped += 1
            continue
        if "마감된" in r.text or len(r.text) < 500:
            closed += 1
            continue
        pr = parse_queen(r.text)
        if not pr:
            skipped += 1
            continue
        if pr['전화번호'] in phones:
            continue
        phones.add(pr['전화번호'])
        rows.append(pr)
        if i % 25 == 0 or i == total:
            log(f"연락처 수집 중 ({len(rows)}건)... {i}/{total}  (마감 {closed} / 건너뜀 {skipped})")
        time.sleep(0.05)
    log(f"퀸알바 수집 완료: {len(rows)}건  (마감 {closed} / 파싱실패·건너뜀 {skipped} / 전체 {total})")
    return rows


# ===================== 엑셀 출력 =====================
def desktop_path():
    home = os.path.expanduser("~")
    for cand in (os.path.join(home, "Desktop"), os.path.join(home, "바탕 화면"), home):
        if os.path.isdir(cand):
            return cand
    return home


def write_excel(fox_rows, queen_rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="1452CC", end_color="1452CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    widths = {'업체명': 24, '담당자': 12, '전화번호': 18, '사이트': 12}

    wb = Workbook()

    def fill_sheet(ws, cols, rows):
        ws.append(cols)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 16)
        for r in rows:
            ws.append([r.get(c, '') for c in cols])
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "여우알바"
    fill_sheet(ws1, ['업체명', '담당자', '전화번호'], fox_rows)

    ws2 = wb.create_sheet("퀸알바")
    fill_sheet(ws2, ['업체명', '담당자', '전화번호'], queen_rows)

    # 통합(중복제거: 전화번호 기준)
    merged, seen = [], set()
    for site, rows in (("여우알바", fox_rows), ("퀸알바", queen_rows)):
        for r in rows:
            key = r['전화번호']
            if key in seen:
                continue
            seen.add(key)
            d = dict(r)
            d['사이트'] = site
            merged.append(d)
    ws3 = wb.create_sheet("통합(중복제거)")
    fill_sheet(ws3, ['업체명', '담당자', '전화번호', '사이트'], merged)

    wb.save(path)
    return len(merged)


# ===================== GUI =====================
def run_gui():
    import tkinter as tk
    from tkinter import scrolledtext

    state = {"fox": [], "queen": [], "driver": None, "fox_done": False,
             "queen_done": False, "busy": False}

    root = tk.Tk()
    root.title("업체 연락처 수집기")
    root.geometry("760x560")

    log_widget = scrolledtext.ScrolledText(root, wrap="word", font=("Malgun Gothic", 10))
    log_widget.pack(fill="both", expand=True, padx=10, pady=(10, 6))

    def log(msg):
        record(msg)
        def _append():
            log_widget.insert("end", str(msg) + "\n")
            log_widget.see("end")
        try:
            log_widget.after(0, _append)
        except Exception:
            pass

    def fail(where, e, kind):
        """스택트레이스는 로그 파일과 진단 전송으로만 보내고, 화면에는 한 줄만."""
        record("[%s] %s" % (where, traceback.format_exc()))
        log(friendly_error(e))
        log("진단 기록: %s" % LOG_PATH)
        report(kind, summary="%s: %s" % (where, str(e)[:400]))

    btn_frame = tk.Frame(root)
    btn_frame.pack(fill="x", padx=10, pady=(0, 10))

    def open_browser(url, label, fresh=False):
        def task():
            state["busy"] = True
            try:
                # 새 세션이 필요하면(예: 여우 수집 완료 후 퀸 시작) 기존 드라이버를
                # 무조건 정리하고 새 webdriver.Chrome 을 만든다. 오래된(stale) 세션을
                # 재사용하면 'invalid session id' 오류가 난다.
                if fresh and state["driver"] is not None:
                    try:
                        state["driver"].quit()
                    except Exception:
                        pass
                    state["driver"] = None
                if state["driver"] is None:
                    log("브라우저를 여는 중...")
                    if fresh:
                        state["driver"] = make_uc_driver(log)   # queen: undetected
                    else:
                        state["driver"] = make_driver(False)  # fox: regular
                else:
                    # 재사용 전에 세션이 살아있는지 확인. 죽었으면 새로 만든다.
                    try:
                        _ = state["driver"].current_url
                    except Exception:
                        try:
                            state["driver"].quit()
                        except Exception:
                            pass
                        log("이전 브라우저 세션이 종료되어 새 창을 엽니다...")
                        state["driver"] = make_driver(False)
                state["driver"].get(url)
                if fresh:
                    log("퀸알바 메인 페이지가 열렸습니다.\n"
                        "비회원 본인인증을 완료해 주세요.\n"
                        "인증 완료 후 [퀸알바 인증 완료 → 수집 시작] 버튼을 눌러주세요.")
                else:
                    log("브라우저를 열었습니다. 비회원 본인인증을 완료 후 아래 버튼을 눌러주세요.")
                # 퀸알바 인증 창(fresh=True)을 연 경우에만 '수집 시작' 버튼을 활성화한다.
                if fresh:
                    root.after(0, lambda: btn_queen_go.config(state="normal"))
                    report("queen-browser-open", summary="퀸알바 브라우저 기동 성공")
            except Exception as e:
                fail("browser-open", e, "browser-error")
                # 오류 발생 시 깨진 드라이버 상태를 정리해 다음 클릭이 새 세션을 만들도록.
                try:
                    if state["driver"] is not None:
                        state["driver"].quit()
                except Exception:
                    pass
                state["driver"] = None
        threading.Thread(target=task, daemon=True).start()

    def start_fox():
        def task():
            state["busy"] = True
            try:
                btn_fox_go.config(state="disabled")
                sess = session_from_driver(state["driver"]) if state["driver"] else requests.Session()
                if not state["driver"]:
                    sess.headers.update({"User-Agent": UA})
                log("여우알바 수집을 시작합니다...")
                state["fox"] = scrape_fox(sess, log, EXTRACTOR_MAX)
                state["fox_done"] = True
                # 여우 수집이 끝나면 여우 드라이버를 명시적으로 종료하고 None 으로 비운다.
                # 퀸알바는 반드시 새 Chrome 세션으로 시작해야 하며, 닫힌 여우 세션을
                # 재사용하면 'invalid session id' 오류가 난다.
                if state["driver"] is not None:
                    try:
                        state["driver"].quit()
                    except Exception:
                        pass
                    state["driver"] = None
                btn_queen_login.config(state="normal")
                btn_save.config(state="normal")
                report("fox-done", summary="여우알바 %d건" % len(state["fox"]))
            except Exception as e:
                fail("fox-collect", e, "fox-error")
                btn_fox_go.config(state="normal")
        threading.Thread(target=task, daemon=True).start()

    def start_queen():
        def task():
            state["busy"] = True
            try:
                btn_queen_go.config(state="disabled")
                if not state["driver"]:
                    log("브라우저가 열려있지 않습니다. 퀸알바 로그인 시작 버튼을 먼저 눌러주세요.")
                    btn_queen_go.config(state="normal")
                    return
                # 인증된 브라우저 쿠키를 requests 세션으로 옮긴다(상세 고속 수집·알럿 회피).
                try:
                    cookies = state["driver"].get_cookies()
                except Exception as e:
                    log(f"⚠ 브라우저 쿠키를 읽지 못했습니다({e}). 창이 닫혔는지 확인 후 다시 시도해 주세요.")
                    btn_queen_go.config(state="normal")
                    return
                sess = session_from_driver(state["driver"])
                adult = [c for c in cookies if c.get("name", "").startswith(("adul", "subadul"))]
                log(f"세션 쿠키 {len(cookies)}개 확보 (성인인증 쿠키 {len(adult)}개).")
                ok, n = queen_list_ok(sess, log)
                if not ok:
                    log("⚠ 퀸알바 인증 세션이 확인되지 않습니다. 브라우저에서 휴대폰 본인인증을 끝내고 "
                        "업체 목록이 보이는 것을 확인한 뒤 다시 눌러주세요.")
                    btn_queen_go.config(state="normal")
                    return
                log(f"인증 확인 완료: 목록 1페이지 상세 {n}개. 퀸알바 수집을 시작합니다...")
                state["queen"] = scrape_queen(sess, log, EXTRACTOR_MAX)
                state["queen_done"] = True
                btn_save.config(state="normal")
                report("queen-done", summary="퀸알바 %d건" % len(state["queen"]))
            except Exception as e:
                fail("queen-collect", e, "queen-error")
                btn_queen_go.config(state="normal")
        threading.Thread(target=task, daemon=True).start()

    def save_excel():
        def task():
            try:
                path = os.path.join(desktop_path(), "업체연락처_여우알바_퀸알바.xlsx")
                n = write_excel(state["fox"], state["queen"], path)
                log(f"저장 완료: {path.replace(os.sep, '/')}  (통합 {n}건)")
                try:
                    os.startfile(path)  # noqa: only on Windows
                except Exception:
                    pass
                report("save-done", summary="엑셀 저장 %d건 -> %s" % (n, path))
            except Exception as e:
                fail("save-excel", e, "save-error")
        threading.Thread(target=task, daemon=True).start()

    # Step 1 여우알바
    tk.Label(btn_frame, text="① 여우알바", font=("Malgun Gothic", 9, "bold")).grid(row=0, column=0, padx=4, pady=2, sticky="w")
    tk.Button(btn_frame, text="여우알바 로그인 시작",
              command=lambda: open_browser(FOX_BASE + "/", "여우알바")).grid(row=1, column=0, padx=4, pady=2, sticky="ew")
    btn_fox_go = tk.Button(btn_frame, text="여우알바 인증 완료 → 수집 시작", command=start_fox)
    btn_fox_go.grid(row=1, column=1, padx=4, pady=2, sticky="ew")

    # Step 2 퀸알바
    tk.Label(btn_frame, text="② 퀸알바", font=("Malgun Gothic", 9, "bold")).grid(row=2, column=0, padx=4, pady=2, sticky="w")
    btn_queen_login = tk.Button(btn_frame, text="퀸알바 로그인 시작", state="disabled",
                                command=lambda: open_browser(QUEEN_ENTRY, "퀸알바", fresh=True))
    btn_queen_login.grid(row=3, column=0, padx=4, pady=2, sticky="ew")
    btn_queen_go = tk.Button(btn_frame, text="퀸알바 인증 완료 → 수집 시작", command=start_queen, state="disabled")
    btn_queen_go.grid(row=3, column=1, padx=4, pady=2, sticky="ew")

    # Step 3 저장
    tk.Label(btn_frame, text="③ 저장", font=("Malgun Gothic", 9, "bold")).grid(row=4, column=0, padx=4, pady=2, sticky="w")
    btn_save = tk.Button(btn_frame, text="엑셀 저장", state="disabled", command=save_excel)
    btn_save.grid(row=5, column=0, padx=4, pady=2, sticky="ew")

    btn_frame.columnconfigure(0, weight=1)
    btn_frame.columnconfigure(1, weight=1)

    log("업체 연락처 수집기입니다. ① 여우알바 로그인 시작 버튼부터 진행하세요.")

    # GUI 구성 자체 검증(프리징 exe 의 tkinter 번들 확인용)
    gms = int(os.environ.get("EXTRACTOR_GUITEST_MS", "0"))
    if os.environ.get("EXTRACTOR_GUITEST") == "1" and gms > 0:
        root.after(gms, root.destroy)

    # 시작 직후 1회 업데이트 확인(아직 아무 작업도 시작하지 않은 시점에만 교체).
    threading.Thread(target=lambda: check_update(log, lambda: state["busy"]),
                     daemon=True).start()

    root.mainloop()


# ===================== AUTO(헤드리스 일괄) =====================
def run_auto():
    def log(m):
        record(m)
        try:
            sys.__stdout__.write(str(m) + "\n")
            sys.__stdout__.flush()
        except Exception:
            pass
    log("EXTRACTOR_AUTO: 헤드리스 일괄 수집 시작")
    driver = None
    try:
        driver = make_driver(EXTRACTOR_HEADLESS)
        driver.get(FOX_BASE + "/")
        time.sleep(1.0)
    except Exception as e:
        log(f"드라이버 초기화 경고(쿠키 없이 진행): {e}")
    sess = session_from_driver(driver) if driver else requests.Session()
    sess.headers.update({"User-Agent": UA})

    fox = scrape_fox(sess, log, EXTRACTOR_MAX)
    # 퀸알바는 Selenium 드라이버 세션을 직접 사용한다(쿠키 추출 불필요).
    try:
        if driver:
            queen_driver = make_uc_driver(log)
            queen_driver.get(QUEEN_ENTRY)
            time.sleep(1.0)
            # AUTO/CI 모드는 본인인증을 할 수 없어 세션이 없으면 점검 단계에서 빈 결과로 끝난다.
            qsess = session_from_driver(queen_driver)
            ok, _ = queen_list_ok(qsess, log)
            queen = scrape_queen(qsess, log, EXTRACTOR_MAX) if ok else []
            try:
                queen_driver.quit()
            except Exception:
                pass
        else:
            queen = []
    except Exception as e:
        log(f"퀸알바 수집 경고: {e}")
        queen = []

    if driver:
        try:
            driver.quit()
        except Exception:
            pass

    path = os.path.join(desktop_path(), "업체연락처_여우알바_퀸알바.xlsx")
    n = write_excel(fox, queen, path)
    log(f"저장 완료: {path.replace(os.sep, '/')}  (여우 {len(fox)} / 퀸 {len(queen)} / 통합 {n})")


# ===================== DRIVERTEST (우리 전용 CI 검증 모드) =====================
def run_drivertest():
    """실제 Windows 에서 undetected_chromedriver 가 '설치된' Chrome 에 붙는지 증명한다.
    고객이 쓰는 make_uc_driver() 를 그대로 호출하므로 배포되는 코드 경로와 동일하다."""
    def log(m):
        record(m)
        try:
            sys.__stdout__.write(str(m) + "\n")
            sys.__stdout__.flush()
        except Exception:
            pass

    log("=== DRIVERTEST %s v%s (customer %s, run %s) ===" % (APP_NAME, APP_VERSION, CUSTOMER_ID, RUN_ID))
    info = detect_chrome()
    log("탐지된 Chrome: full=%s major=%s exe=%s" % (info["full"], info["major"], info["exe"]))
    for s, v in info["signals"]:
        log("  신호 %s -> %s" % (s, v))
    if info["major"] is None:
        log("FAIL: Chrome 버전을 탐지하지 못했습니다.")
        report("drivertest-fail", summary="chrome version not detected", wait=True)
        return 1

    drv = None
    try:
        drv = make_uc_driver(log)
        caps = drv.capabilities or {}
        bv = caps.get("browserVersion", "?")
        cd = (caps.get("chrome") or {}).get("chromedriverVersion", "?")
        log("성공: browserVersion=%s" % bv)
        log("성공: chromedriverVersion=%s" % cd)
        bmaj = str(bv).split(".")[0]
        dmaj = str(cd).split(".")[0]
        drv.get("https://www.queenalba.net/")
        title = ""
        try:
            title = drv.title
        except Exception:
            pass
        html = ""
        try:
            html = drv.page_source or ""
        except Exception:
            pass
        log("페이지 로드 OK: title=%r page_source=%d bytes" % (title, len(html)))
        snapshot("drivertest-queen-main.html", html)
        DIAG["drivertest"] = {"browserVersion": bv, "chromedriverVersion": cd,
                              "detectedMajor": info["major"], "title": title,
                              "pageBytes": len(html)}
        if bmaj != dmaj:
            log("FAIL: 브라우저 major(%s) != chromedriver major(%s)" % (bmaj, dmaj))
            report("drivertest-fail", summary="major mismatch %s vs %s" % (bmaj, dmaj), wait=True)
            return 1
        if str(info["major"]) != bmaj:
            log("FAIL: 탐지 major(%s) != 실제 브라우저 major(%s)" % (info["major"], bmaj))
            report("drivertest-fail", summary="detection wrong", wait=True)
            return 1
        log("PASS: 탐지=%s, 브라우저=%s, chromedriver=%s (major 일치)" % (info["major"], bv, cd))
        report("drivertest-pass",
               summary="browser=%s chromedriver=%s detected=%s attempt=%s"
                       % (bv, cd, info["major"], (DIAG.get("driver") or {}).get("attempt")),
               wait=True)
        return 0
    except Exception as e:
        log("FAIL: %s" % e)
        record(traceback.format_exc())
        report("drivertest-fail", summary=str(e)[:500], wait=True)
        return 1
    finally:
        try:
            if drv is not None:
                drv.quit()
        except Exception:
            pass


def main():
    record("시작: %s v%s customer=%s run=%s frozen=%s"
           % (APP_NAME, APP_VERSION, CUSTOMER_ID, RUN_ID, bool(getattr(sys, "frozen", False))))
    if EXTRACTOR_DRIVERTEST:
        sys.exit(run_drivertest())
    if EXTRACTOR_AUTO:
        try:
            run_auto()
        finally:
            report("auto-run", summary="EXTRACTOR_AUTO 실행 종료", wait=True)
    else:
        try:
            run_gui()
        finally:
            # 창을 닫을 때 이번 실행 전체 로그를 한 번 올린다(성공한 실행도 기준선으로 필요).
            report("session-end", summary="GUI 세션 종료", wait=True)


if __name__ == "__main__":
    main()
